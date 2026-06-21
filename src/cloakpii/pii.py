"""
PII detection and desensitization for multiple file formats.

Supported PII types:
  - email        user@example.com  →  u***@e******.com
  - phone        555-123-4567      →  555-***-****
  - national_id  123-45-6789 (SSN) →  ***-**-6789
  - credit_card  4111111111111111  →  4111****1111
  - ip_address   192.168.1.100     →  192.168.*.*
  - chinese_id   110101199001011234 →  1101***********234
  - passport     AB1234567         →  AB***4567
  - bank_account 1234567890123456  →  1234********3456
  - iban         GB29NWBK60161331926819 →  GB29****6819
  - mac_address  00:1B:44:11:3A:B7 →  00:1B:**:**:**:B7

Supported file formats:
  CSV, JSON, Excel (.xlsx), Parquet (.parquet), XML, TSV, SQLite

Field-name heuristics (for column names):
  If a column name contains keywords like 'name', 'email', 'phone', 'ssn',
  'address', 'id_number', the value is masked regardless of content pattern.
"""

import csv
import json
import re
import sqlite3
import shutil
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    import defusedxml.ElementTree as DET
except ImportError:
    DET = None


# ---------------------------------------------------------------------------
# PII patterns
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
# Phone numbers must carry a signal that they ARE phone numbers — either an
# international "+CC" prefix or at least one separator between digit groups.
# This avoids masking bare integers (order IDs, counts, timestamps), which the
# old greedy pattern swept up as false positives.
PHONE_RE = re.compile(
    r"\+\d{1,3}[-.\s]?\d{2,4}[-.\s]?\d{3,4}[-.\s]?\d{3,4}"  # +CC international
    r"|"
    r"\(?\d{2,4}\)?[-.\s]\d{3,4}[-.\s]?\d{3,4}"            # domestic, ≥1 separator
)
SSN_RE = re.compile(r"\b\d{3}-\d{2}-\d{4}\b")
CREDIT_CARD_RE = re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b")
IP_RE = re.compile(r"\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b")

# New PII patterns
CHINESE_ID_RE = re.compile(r"\b\d{17}[\dXx]\b")
PASSPORT_RE = re.compile(r"\b[A-Z]{1,2}\d{6,9}\b")
BANK_ACCOUNT_RE = re.compile(r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4,7}\b")
DATE_OF_BIRTH_RE = re.compile(
    r"\b(?:19|20)\d{2}[-/](?:0[1-9]|1[0-2])[-/](?:0[1-9]|[12]\d|3[01])\b|"
    r"\b(?:0[1-9]|1[0-2])[-/](?:0[1-9]|[12]\d|3[01])[-/](?:19|20)\d{2}\b"
)
IBAN_RE = re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{4}\d{7,16}\b")
MAC_ADDRESS_RE = re.compile(r"\b(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}\b")

# Column-name keywords that trigger name-based masking
NAME_KEYWORDS = {
    # English keywords
    "name", "email", "phone", "ssn", "address", "id_number",
    "national_id", "credit_card", "fullname", "full_name",
    "mobile", "tel", "telephone", "passport", "bank_account",
    "iban", "date_of_birth", "dob", "chinese_id", "mac_address",
    "license_number", "driver_license", "account_number",
    # Chinese keywords (重要：针对亚洲市场)
    "姓名", "邮箱", "电子邮箱", "手机", "手机号", "电话", "电话号码",
    "身份证", "身份证号", "身份证号码", "地址", "住址", "银行账户",
    "银行卡", "银行卡号", "护照", "护照号", "信用卡", "信用卡号",
    "出生日期", "生日", "mac地址", "驾照", "驾照号",
}


# ---------------------------------------------------------------------------
# Masking helpers
# ---------------------------------------------------------------------------

def mask_email(val: str) -> str:
    """user@example.com → u***@e******.com"""
    m = EMAIL_RE.search(val)
    if not m:
        return val
    local, domain = m.group().split("@", 1)
    masked_local = local[0] + "***" if len(local) > 1 else "***"
    dname, tld = domain.rsplit(".", 1) if "." in domain else (domain, "")
    masked_domain = dname[0] + "******" if dname else "******"
    result = f"{masked_local}@{masked_domain}.{tld}"
    return val[: m.start()] + result + val[m.end():]


def mask_phone(val: str) -> str:
    """Mask phone number, keeping country code (if any) and last 2 digits visible.
    
    Examples:
        555-123-4567 → 555-***-**67
        +1-555-123-4567 → +1-***-***-**67
        +86-138-1234-5678 → +86-***-****-**78
        416-555-0123 → 416-***-**23
    """
    digits = re.findall(r"\d", val)
    if len(digits) < 7:
        return val
    
    # Detect if there's a country code prefix (+ or 00)
    has_plus = val.strip().startswith("+")
    
    # Keep last 2 digits visible, mask the rest
    if has_plus:
        # For international: keep +CC and last 2
        # Find where country code ends (after first group of 1-3 digits following +)
        cc_match = re.match(r"(\+\d{1,3}[-.\s]?)", val)
        if cc_match:
            prefix = cc_match.group(1)
            rest = val[len(prefix):]
            # Mask all digits in rest except last 2
            result = []
            digit_count = len(re.findall(r"\d", rest))
            digits_to_mask = digit_count - 2
            masked_so_far = 0
            for ch in rest:
                if ch.isdigit():
                    if masked_so_far < digits_to_mask:
                        result.append("*")
                        masked_so_far += 1
                    else:
                        result.append(ch)
                else:
                    result.append(ch)
            return prefix + "".join(result)
    
    # For domestic: keep first 3 digits and last 2
    result = []
    digit_idx = 0
    total_digits = len(digits)
    for ch in val:
        if ch.isdigit():
            if digit_idx < 3 or digit_idx >= total_digits - 2:
                result.append(ch)
            else:
                result.append("*")
            digit_idx += 1
        else:
            result.append(ch)
    return "".join(result)


def mask_ssn(val: str) -> str:
    return "***-**-" + val[-4:]


def luhn_valid(number: str) -> bool:
    """Return True if the digit string passes the Luhn checksum (all card
    networks use it). Filters out random numeric strings that merely look
    card-shaped."""
    digits = [int(c) for c in number if c.isdigit()]
    if len(digits) < 12:
        return False
    checksum = 0
    parity = len(digits) % 2
    for i, d in enumerate(digits):
        if i % 2 == parity:
            d *= 2
            if d > 9:
                d -= 9
        checksum += d
    return checksum % 10 == 0


def mask_credit_card(val: str) -> str:
    clean = re.sub(r"[\s-]", "", val)
    if len(clean) < 8:
        return val
    # Only treat as a credit card if it passes the Luhn checksum; otherwise
    # leave it for a more specific matcher (e.g. bank account) to handle.
    if not luhn_valid(clean):
        return val
    return clean[:4] + "****" + clean[-4:]


def mask_ip(val: str) -> str:
    parts = val.split(".")
    if len(parts) != 4:
        return val
    return f"{parts[0]}.{parts[1]}.*.*"


def mask_generic(val: str) -> str:
    """Generic masking: keep first char, replace rest with ***."""
    if not val or len(val) < 2:
        return "***"
    return val[0] + "***"


def mask_chinese_id(val: str) -> str:
    """110101199001011234 → 1101***********234"""
    if len(val) < 8:
        return val
    return val[:4] + "*" * (len(val) - 7) + val[-3:]


def mask_passport(val: str) -> str:
    """AB1234567 → AB***4567"""
    if len(val) < 5:
        return val
    return val[:2] + "***" + val[-4:]


def mask_bank_account(val: str) -> str:
    """1234567890123456 → 1234********3456"""
    clean = re.sub(r"[\s-]", "", val)
    if len(clean) < 8:
        return val
    return clean[:4] + "*" * (len(clean) - 8) + clean[-4:]


def mask_dob(val: str) -> str:
    """1990-01-15 → ****-**-15"""
    # Keep last 2 digits (day) visible
    parts = re.split(r"[-/]", val)
    if len(parts) == 3:
        return "****-**-" + parts[-1]
    return val


def mask_iban(val: str) -> str:
    """GB29NWBK60161331926819 → GB29****6819"""
    if len(val) < 8:
        return val
    return val[:4] + "****" + val[-4:]


def mask_mac(val: str) -> str:
    """00:1B:44:11:3A:B7 → 00:1B:**:**:**:B7"""
    parts = re.split(r"[:-]", val)
    if len(parts) != 6:
        return val
    return f"{parts[0]}:{parts[1]}:**:**:**:{parts[5]}"


# ---------------------------------------------------------------------------
# Auto-detect & mask a single string value
# ---------------------------------------------------------------------------

def mask_value(val: str) -> str:
    """Apply all regex-based PII masks to a string value."""
    if not isinstance(val, str):
        return val
    result = val
    # Order matters: more specific patterns first
    result = SSN_RE.sub(lambda m: mask_ssn(m.group()), result)
    result = CHINESE_ID_RE.sub(lambda m: mask_chinese_id(m.group()), result)
    result = IBAN_RE.sub(lambda m: mask_iban(m.group()), result)
    result = CREDIT_CARD_RE.sub(lambda m: mask_credit_card(m.group()), result)
    result = BANK_ACCOUNT_RE.sub(lambda m: mask_bank_account(m.group()), result)
    result = PASSPORT_RE.sub(lambda m: mask_passport(m.group()), result)
    result = EMAIL_RE.sub(lambda m: mask_email(m.group()), result)
    result = IP_RE.sub(lambda m: mask_ip(m.group()), result)
    result = MAC_ADDRESS_RE.sub(lambda m: mask_mac(m.group()), result)
    result = PHONE_RE.sub(lambda m: mask_phone(m.group()), result)
    return result


def _is_pii_field(field_name: str) -> bool:
    """Check if a column/field name suggests PII content."""
    lower = field_name.lower().replace(" ", "_").replace("-", "_")
    return any(kw in lower for kw in NAME_KEYWORDS)


# Per-field policy actions. A field may be explicitly assigned one of these,
# overriding the global mode and the auto-detector.
FIELD_ACTIONS = {"mask", "tokenize", "drop", "keep"}


def resolve_field_action(field_name, field_policies):
    """Return the explicit policy action for a field, or None for default behaviour.

    ``field_policies`` is a dict of normalized (lower-cased) field name → action.
    """
    if not field_policies or not field_name:
        return None
    return field_policies.get(str(field_name).strip().lower())


def _dropped_fields(names, field_policies):
    """Return the set of field names whose policy action is ``drop``."""
    if not field_policies:
        return set()
    return {n for n in names if resolve_field_action(n, field_policies) == "drop"}


def _transform_cell(value, field_name="", mode="mask", tokenizer=None, field_policies=None):
    """Single source of truth for transforming one cell/value.

    Modes:
      - "mask"       : irreversible partial masking (default, original behaviour)
      - "tokenize"   : replace a whole PII value with a stable reversible token
      - "detokenize" : reverse any tokens found back to the original value

    ``field_policies`` (optional) maps a field name to an explicit action
    (mask/tokenize/keep/drop) that overrides the global mode and auto-detection.
    ``drop`` is structural (the column/key is removed by the caller); here it
    just leaves the value untouched.

    Returns (new_value, changed).
    """
    # detokenize restores originals — field policies don't apply.
    if mode == "detokenize":
        if not isinstance(value, str) or not value:
            return value, False
        from .tokenize import TOKEN_RE
        if TOKEN_RE.search(value):
            return tokenizer.detokenize_text(value), True
        return value, False

    action = resolve_field_action(field_name, field_policies)
    if action in ("keep", "drop"):
        # keep: leave untouched. drop: the caller removes the field entirely.
        return value, False

    if not isinstance(value, str) or not value:
        return value, False

    # Explicit per-field action wins over the global mode / auto-detector.
    if action == "tokenize":
        if not value.strip():
            return value, False
        return tokenizer.tokenize(value), True
    if action == "mask":
        masked = mask_value(value)
        if masked == value and value.strip():
            masked = mask_generic(value)  # forced — the user pinned this field to mask
        return masked, masked != value

    # No explicit policy → original auto behaviour based on the global mode.
    if mode == "tokenize":
        if not value.strip():
            return value, False
        is_pii = (bool(field_name) and _is_pii_field(field_name)) or (mask_value(value) != value)
        if is_pii:
            return tokenizer.tokenize(value), True
        return value, False

    # default: mask
    masked = mask_value(value)
    if masked == value and field_name and _is_pii_field(field_name) and value.strip():
        masked = mask_generic(value)
    return masked, masked != value


# ---------------------------------------------------------------------------
# Desensitization stats
# ---------------------------------------------------------------------------

@dataclass
class DesensitizeReport:
    """Tracks what was masked during a desensitization pass."""
    fields_masked: list = field(default_factory=list)
    values_masked: int = 0
    rows_processed: int = 0


# ---------------------------------------------------------------------------
# Text
# ---------------------------------------------------------------------------

def desensitize_text(text: str, mode="mask", tokenizer=None) -> tuple:
    """Transform PII in plain text. Returns (new_text, count_of_replacements).

    mode="mask" partially masks each match; mode="tokenize" replaces each match
    with a reversible token; mode="detokenize" restores tokens to originals.
    """
    if mode == "detokenize":
        from .tokenize import TOKEN_RE
        count = len(TOKEN_RE.findall(text))
        return tokenizer.detokenize_text(text), count

    count = 0

    # Define all pattern-masker pairs explicitly
    pattern_maskers = [
        (SSN_RE, mask_ssn),
        (CHINESE_ID_RE, mask_chinese_id),
        (IBAN_RE, mask_iban),
        (CREDIT_CARD_RE, mask_credit_card),
        (BANK_ACCOUNT_RE, mask_bank_account),
        (PASSPORT_RE, mask_passport),
        (EMAIL_RE, mask_email),
        (IP_RE, mask_ip),
        (MAC_ADDRESS_RE, mask_mac),
        (PHONE_RE, mask_phone),
    ]

    # Process each pattern
    for pattern, masker_fn in pattern_maskers:
        def make_replacer(fn):
            """Create a replacer function with proper closure."""
            def _replace(m):
                nonlocal count
                count += 1
                if mode == "tokenize":
                    return tokenizer.tokenize(m.group())
                return fn(m.group())
            return _replace

        text = pattern.sub(make_replacer(masker_fn), text)

    # Apply custom patterns
    for name, pattern in CUSTOM_PII_PATTERNS:
        try:
            custom_re = re.compile(pattern)
            def make_custom_replacer(pname):
                def _replace(m):
                    nonlocal count
                    count += 1
                    if mode == "tokenize":
                        return tokenizer.tokenize(m.group())
                    return f"[{pname}]"
                return _replace
            text = custom_re.sub(make_custom_replacer(name), text)
        except re.error:
            continue

    return text, count


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def desensitize_csv(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                    field_policies=None) -> DesensitizeReport:
    """Read CSV, transform PII in every cell, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        dropped = _dropped_fields(fieldnames, field_policies)
        out_fields = [fn for fn in fieldnames if fn not in dropped]

        # Determine which fields need name-based masking
        pii_fields = [fn for fn in out_fields if _is_pii_field(fn)]
        report.fields_masked = list(pii_fields)

        rows = []
        for row in reader:
            report.rows_processed += 1
            out_row = {}
            for fn in out_fields:
                original = row.get(fn, "")
                new_value, changed = _transform_cell(original, fn, mode, tokenizer, field_policies)
                if changed:
                    report.values_masked += 1
                out_row[fn] = new_value
            rows.append(out_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields)
        writer.writeheader()
        writer.writerows(rows)

    return report


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def _desensitize_json_node(node: Any, report: DesensitizeReport, parent_key: str = "",
                           mode="mask", tokenizer=None, field_policies=None) -> Any:
    """Recursively walk a JSON structure and transform PII values."""
    if isinstance(node, dict):
        result = {}
        for k, v in node.items():
            # Drop a key entirely if its policy says so (mask/tokenize runs only).
            if mode != "detokenize" and resolve_field_action(k, field_policies) == "drop":
                continue
            result[k] = _desensitize_json_node(v, report, parent_key=k, mode=mode,
                                               tokenizer=tokenizer, field_policies=field_policies)
        return result
    elif isinstance(node, list):
        return [_desensitize_json_node(item, report, parent_key=parent_key, mode=mode,
                                       tokenizer=tokenizer, field_policies=field_policies)
                for item in node]
    elif isinstance(node, str):
        new_value, changed = _transform_cell(node, parent_key, mode, tokenizer, field_policies)
        if changed:
            report.values_masked += 1
            if parent_key and parent_key not in report.fields_masked:
                report.fields_masked.append(parent_key)
        return new_value
    elif isinstance(node, bool):
        # bool is a subclass of int — never treat True/False as PII.
        return node
    elif isinstance(node, (int, float)):
        # Numeric PII (phones, IDs, account numbers stored as numbers) would
        # otherwise pass through unmasked. Mask on the string form; if it
        # changes, emit the masked string (the value is no longer a number).
        new_value, changed = _transform_cell(str(node), parent_key, mode, tokenizer, field_policies)
        if changed:
            report.values_masked += 1
            if parent_key and parent_key not in report.fields_masked:
                report.fields_masked.append(parent_key)
            return new_value
        return node
    else:
        return node


def desensitize_json(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                     field_policies=None) -> DesensitizeReport:
    """Read JSON, transform PII recursively, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        report.rows_processed = len(data)

    masked_data = _desensitize_json_node(data, report, mode=mode, tokenizer=tokenizer,
                                         field_policies=field_policies)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(masked_data, f, indent=2, ensure_ascii=False)

    return report


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------

def desensitize_excel(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                      field_policies=None) -> DesensitizeReport:
    """Read Excel workbook, transform PII in all sheets, write to output_path."""
    import openpyxl

    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(input_path, read_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue

        # First row is header
        header_cells = rows[0]
        fieldnames = [cell.value or "" for cell in header_cells]

        # 1-based column indices to drop (only in mask/tokenize runs).
        drop_idx = []
        if mode != "detokenize" and field_policies:
            drop_idx = [i + 1 for i, fn in enumerate(fieldnames)
                        if resolve_field_action(str(fn), field_policies) == "drop"]

        pii_fields = [fn for i, fn in enumerate(fieldnames)
                      if (i + 1) not in drop_idx and _is_pii_field(str(fn))]
        for pf in pii_fields:
            if pf not in report.fields_masked:
                report.fields_masked.append(pf)

        for row in rows[1:]:
            report.rows_processed += 1
            for idx, cell in enumerate(row):
                if (idx + 1) in drop_idx or cell.value is None:
                    continue
                original = str(cell.value)
                fn = fieldnames[idx] if idx < len(fieldnames) else ""
                new_value, changed = _transform_cell(original, str(fn), mode, tokenizer, field_policies)
                if changed:
                    report.values_masked += 1
                    cell.value = new_value

        # Delete dropped columns right-to-left so indices stay valid.
        for col in sorted(drop_idx, reverse=True):
            ws.delete_cols(col)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return report


# ---------------------------------------------------------------------------
# Parquet (.parquet)
# ---------------------------------------------------------------------------

def desensitize_parquet(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                        field_policies=None) -> DesensitizeReport:
    """Read Parquet file, transform PII in string columns, write to output_path."""
    import pyarrow.parquet as pq
    import pyarrow as pa

    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    table = pq.read_table(input_path)
    schema = table.schema

    drop = set()
    if mode != "detokenize" and field_policies:
        drop = {fld.name for fld in schema
                if resolve_field_action(fld.name, field_policies) == "drop"}

    # Identify string columns and PII field names
    pii_fields = []
    string_cols = []
    for i, fld in enumerate(schema):
        if pa.types.is_string(fld.type) or pa.types.is_large_string(fld.type):
            string_cols.append(i)
            if fld.name not in drop and _is_pii_field(fld.name):
                pii_fields.append(fld.name)

    report.fields_masked = list(pii_fields)
    report.rows_processed = table.num_rows

    # Process each column. String columns are masked in place; non-string
    # columns (ints/floats that may hold phones, IDs, account numbers) are
    # scanned too — if any value is PII the whole column is emitted as strings,
    # since a masked value (e.g. "138-****-**78") is no longer numeric.
    new_columns = []
    out_names = []
    for i, fld in enumerate(schema):
        if fld.name in drop:
            continue  # column removed from output
        out_names.append(fld.name)
        col = table.column(i)
        if i in string_cols:
            masked_values = []
            for val in col.to_pylist():
                if val is None:
                    masked_values.append(None)
                    continue
                original = str(val)
                new_value, changed = _transform_cell(original, fld.name, mode, tokenizer, field_policies)
                if changed:
                    report.values_masked += 1
                masked_values.append(new_value)
            new_columns.append(pa.array(masked_values, type=fld.type))
        elif pa.types.is_nested(fld.type) or pa.types.is_binary(fld.type):
            # Structs/lists/maps/binary aren't scalar PII text — leave as-is.
            new_columns.append(col)
        else:
            masked_values = []
            col_changed = False
            for val in col.to_pylist():
                if val is None:
                    masked_values.append(None)
                    continue
                new_value, changed = _transform_cell(str(val), fld.name, mode, tokenizer, field_policies)
                if changed:
                    report.values_masked += 1
                    col_changed = True
                    masked_values.append(new_value)
                else:
                    masked_values.append(val)
            if col_changed:
                if fld.name not in report.fields_masked:
                    report.fields_masked.append(fld.name)
                # Whole column becomes string; stringify the untouched values too.
                masked_values = [None if v is None else str(v) for v in masked_values]
                new_columns.append(pa.array(masked_values, type=pa.string()))
            else:
                new_columns.append(col)

    new_table = pa.table(dict(zip(out_names, new_columns)))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(new_table, output_path)

    return report


# ---------------------------------------------------------------------------
# XML (.xml)
# ---------------------------------------------------------------------------

def desensitize_xml(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                    field_policies=None) -> DesensitizeReport:
    """Read XML file, transform PII in text content and attributes, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    if DET is not None:
        tree = DET.parse(input_path)
    else:
        # ponytail: no defusedxml — strip DOCTYPE to block XXE
        raw = input_path.read_text(encoding="utf-8")
        safe = re.sub(r"<!DOCTYPE[^>]*\[.*?\]>", "", raw, flags=re.DOTALL)
        safe = re.sub(r"<!DOCTYPE[^>]*>", "", safe)
        tree = ET.ElementTree(ET.fromstring(safe))
    root = tree.getroot()
    drops = mode != "detokenize" and bool(field_policies)

    def _mask_element(elem):
        """Recursively transform PII in an XML element."""
        # Drop child elements whose tag policy says so (mask/tokenize runs only).
        if drops:
            for child in list(elem):
                if resolve_field_action(child.tag, field_policies) == "drop":
                    elem.remove(child)

        # Element text — use the tag name as the field-name hint
        if elem.text and elem.text.strip():
            new_value, changed = _transform_cell(elem.text, elem.tag, mode, tokenizer, field_policies)
            if changed:
                report.values_masked += 1
                if _is_pii_field(elem.tag) and elem.tag not in report.fields_masked:
                    report.fields_masked.append(elem.tag)
                elem.text = new_value

        # Tail text (after closing tag) — no field-name context
        if elem.tail and elem.tail.strip():
            new_value, changed = _transform_cell(elem.tail, "", mode, tokenizer, field_policies)
            if changed:
                report.values_masked += 1
                elem.tail = new_value

        # Attributes
        for attr_name, attr_val in list(elem.attrib.items()):
            if drops and resolve_field_action(attr_name, field_policies) == "drop":
                del elem.attrib[attr_name]
                continue
            new_value, changed = _transform_cell(attr_val, attr_name, mode, tokenizer, field_policies)
            if changed:
                report.values_masked += 1
                if attr_name not in report.fields_masked:
                    report.fields_masked.append(attr_name)
                elem.set(attr_name, new_value)

        report.rows_processed += 1

        # Recurse into children
        for child in elem:
            _mask_element(child)

    _mask_element(root)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Add indentation for readable output (Python 3.9+)
    try:
        ET.indent(tree, space="  ")
    except AttributeError:
        pass  # ET.indent not available in Python < 3.9
    tree.write(output_path, encoding="unicode", xml_declaration=True)

    return report


# ---------------------------------------------------------------------------
# TSV (Tab-Separated Values)
# ---------------------------------------------------------------------------

def desensitize_tsv(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                    field_policies=None) -> DesensitizeReport:
    """Read TSV file, transform PII in every cell, write to output_path."""
    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    with open(input_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = reader.fieldnames or []

        dropped = _dropped_fields(fieldnames, field_policies)
        out_fields = [fn for fn in fieldnames if fn not in dropped]

        # Determine which fields need name-based masking
        pii_fields = [fn for fn in out_fields if _is_pii_field(fn)]
        report.fields_masked = list(pii_fields)

        rows = []
        for row in reader:
            report.rows_processed += 1
            out_row = {}
            for fn in out_fields:
                original = row.get(fn, "")
                new_value, changed = _transform_cell(original, fn, mode, tokenizer, field_policies)
                if changed:
                    report.values_masked += 1
                out_row[fn] = new_value
            rows.append(out_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)

    return report


# ---------------------------------------------------------------------------
# SQLite (.db, .sqlite, .sqlite3)
# ---------------------------------------------------------------------------

def desensitize_sqlite(input_path: Path, output_path: Path, mode="mask", tokenizer=None,
                       field_policies=None) -> DesensitizeReport:
    """Copy a SQLite database and transform PII in every non-BLOB column.

    Masking runs on a temporary copy that is promoted to ``output_path`` only
    after it completes — so a mid-masking failure never leaves an unmasked (or
    partially-masked) database behind. Rows are identified by ``rowid`` where
    available, or by primary key for ``WITHOUT ROWID`` tables.
    """
    import os
    import tempfile

    report = DesensitizeReport()
    input_path = Path(input_path)
    output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(dir=output_path.parent, suffix=".db.tmp")
    os.close(fd)
    tmp_path = Path(tmp_name)

    try:
        shutil.copy(input_path, tmp_path)

        with sqlite3.connect(tmp_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]

            for table_name in tables:
                # Validate table name to prevent SQL injection
                if not table_name.isidentifier():
                    continue

                cursor.execute(f'PRAGMA table_info("{table_name}");')
                columns = cursor.fetchall()
                # columns: (cid, name, type, notnull, dflt_value, pk)
                col_names = [col[1] for col in columns]

                # Columns to drop entirely (mask/tokenize runs only).
                drop_cols = []
                if mode != "detokenize" and field_policies:
                    drop_cols = [c for c in col_names
                                 if resolve_field_action(c, field_policies) == "drop"]

                # Mask every column except BLOBs (and ones we're about to drop).
                # SQLite is dynamically typed, so a masked string can be stored
                # back into an INTEGER/REAL column — this catches PII (phones,
                # IDs, account numbers) held as numbers that a TEXT-only filter
                # would silently leak.
                maskable_cols = [col[1] for col in columns
                                 if "BLOB" not in (col[2] or "").upper()
                                 and col[1] not in drop_cols]
                if not maskable_cols and not drop_cols:
                    continue

                # --- mask remaining columns ---
                if maskable_cols:
                    # Row identification: rowid if the table has one, else primary
                    # key (WITHOUT ROWID tables have no rowid column).
                    try:
                        cursor.execute(f'SELECT rowid FROM "{table_name}" LIMIT 1').fetchall()
                        has_rowid = True
                    except sqlite3.OperationalError:
                        has_rowid = False
                    pk_cols = [col[1] for col in columns if col[5]]  # col[5] = pk position

                    fetched = None
                    if has_rowid:
                        cursor.execute(f'SELECT rowid, * FROM "{table_name}";')
                        fetched = [(r[0], r[1:]) for r in cursor.fetchall()]
                    elif pk_cols:
                        cursor.execute(f'SELECT * FROM "{table_name}";')
                        fetched = [(None, r) for r in cursor.fetchall()]
                    else:
                        import logging
                        logging.getLogger("CloakPII").warning(
                            f"Table '{table_name}' has neither rowid nor a primary "
                            "key; skipping masking (cannot identify rows safely)."
                        )

                    if fetched is not None:
                        report.rows_processed += len(fetched)
                        for col_name in maskable_cols:
                            if _is_pii_field(col_name) and col_name not in report.fields_masked:
                                report.fields_masked.append(col_name)

                        for rowid, row_data in fetched:
                            updates = {}
                            for col_idx, col_name in enumerate(col_names):
                                if col_name not in maskable_cols:
                                    continue
                                original = row_data[col_idx]
                                # None and BLOB (bytes) are not scalar PII text.
                                if original is None or isinstance(original, bytes):
                                    continue
                                new_value, changed = _transform_cell(str(original), col_name, mode, tokenizer, field_policies)
                                if changed:
                                    report.values_masked += 1
                                    updates[col_name] = new_value

                            if not updates:
                                continue

                            set_clause = ", ".join(f'"{k}" = ?' for k in updates)
                            values = list(updates.values())
                            if has_rowid:
                                where_sql, where_params = "rowid = ?", [rowid]
                            else:
                                where_sql = " AND ".join(f'"{pk}" = ?' for pk in pk_cols)
                                where_params = [row_data[col_names.index(pk)] for pk in pk_cols]
                            cursor.execute(
                                f'UPDATE "{table_name}" SET {set_clause} WHERE {where_sql}',
                                values + where_params,
                            )

                # --- drop columns (field policy = drop) ---
                for c in drop_cols:
                    cc = c.replace('"', '""')  # escape identifier quote
                    try:
                        cursor.execute(f'ALTER TABLE "{table_name}" DROP COLUMN "{cc}"')
                    except sqlite3.OperationalError:
                        # Old SQLite (no DROP COLUMN) or indexed/PK column —
                        # null the data out so it is still removed.
                        import logging
                        try:
                            cursor.execute(f'UPDATE "{table_name}" SET "{cc}" = NULL')
                            logging.getLogger("CloakPII").warning(
                                f"Could not drop column '{c}' in '{table_name}'; "
                                "nulled its values instead."
                            )
                        except sqlite3.OperationalError:
                            logging.getLogger("CloakPII").warning(
                                f"Could not drop or null column '{c}' in '{table_name}'."
                            )

            conn.commit()

        os.replace(tmp_path, output_path)  # atomic promote, only on success
    except BaseException:
        # Never leave an unmasked / partially-masked database at output_path.
        try:
            tmp_path.unlink()
        except OSError:
            pass
        raise

    return report

# Custom PII pattern support (v1.2)
CUSTOM_PII_PATTERNS: list[tuple[str, str]] = []

# Classic catastrophic-backtracking shapes: a quantified group whose body is
# itself quantified, e.g. (a+)+, (a*)*, (\d+)*. These can blow up to
# exponential time on adversarial input (ReDoS).
_REDOS_RE = re.compile(r"\([^)]*[+*][^)]*\)\s*[+*]")


def register_custom_pii_pattern(name: str, pattern: str):
    """Register a custom PII detection pattern.

    Patterns come from the user's own config and run against every scanned
    value, so a poorly-written one can cause catastrophic backtracking (ReDoS).
    We warn on the obvious nested-quantifier shape but still register it — the
    user owns the trade-off.
    """
    import re
    try:
        re.compile(pattern)
    except re.error:
        raise ValueError(f"Invalid regex pattern for {name}: {pattern}")
    if _REDOS_RE.search(pattern):
        import logging
        logging.getLogger("CloakPII").warning(
            f"Custom PII pattern '{name}' contains a nested quantifier and may "
            f"be vulnerable to catastrophic backtracking (ReDoS): {pattern}"
        )
    CUSTOM_PII_PATTERNS.append((name, pattern))

def _apply_custom_patterns(text: str) -> list[tuple[str, str]]:
    """Apply registered custom PII patterns to text."""
    import re
    matches = []
    for name, pattern in CUSTOM_PII_PATTERNS:
        try:
            for m in re.finditer(pattern, text):
                matches.append((name, m.group()))
        except re.error:
            continue
    return matches
